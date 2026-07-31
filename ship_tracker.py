#!/usr/bin/env python3
"""
船舶动态跟踪系统 —— 支持 GitHub Actions 定时调度 + 本地运行双模式
========================================================================
使用方式:
  python ship_tracker.py           # 单次查询（GitHub Actions 默认用此模式）
  python ship_tracker.py --daemon  # 本地守护进程模式（每2小时自动查）
  python ship_tracker.py --grab-token  # 本地抓取新 Token 并保存
"""

import os
import sys
import time
import random
import json
import subprocess
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================
# 🔗 可修改链接区 — 你需要修改的链接都在这里
# ==========================================

# 【API 接口地址】NPEDI 船舶排期查询接口
API_URL = "https://www.npedi.com/onesite-api/vessel/plan/selectContainerDynamicPlan"

# 【网页地址】浏览器抓取 Token 时打开的页面
TOKEN_PAGE_URL = "https://www.npedi.com/onesite/vessel/plan"

# 【钉钉 Webhook】（也可通过命令行 --webhook 修改）
DINGTALK_WEBHOOK = "https://oapi.dingtalk.com/robot/send?access_token=53b1f6ace97047e1ffeb5f86a5a84fd8de1bf60beeea4a3ec410435c3b988193"

# ==========================================
# ⚙️ 配置区
# ==========================================
# 路径处理：所有文件都基于脚本所在目录，保证在任何位置运行都能找到
# ==========================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def _path(name):
    return os.path.join(SCRIPT_DIR, name)


TOKEN_FILE = _path("token.txt")
EXCEL_FILE = _path("船舶动态跟踪.xlsx")
STATE_FILE = _path(".last_vessel_state.json")  # 上次运行数据缓存（用于判断数据是否变化）
VESSEL_FILE = _path(".vessels.json")           # 船舶列表文件
DISABLED_FILE = _path(".disabled")             # 关闭查询的标记文件
WEBHOOK_FILE = _path(".webhook.txt")           # 钉钉 Webhook 覆盖文件
UPDATE_INTERVAL_HOURS = 2  # 自动更新间隔（小时）

# ==========================================
# 船舶列表管理（支持命令行修改）
# ==========================================

_DEFAULT_VESSELS = [
    {"name": "YM TOPMOST", "voyage": "025W"},
    {"name": "MAREN MAERSK", "voyage": "630W"},
]


def load_vessels():
    """从文件加载船舶列表，不存在则用默认值"""
    if os.path.exists(VESSEL_FILE):
        try:
            with open(VESSEL_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # 首次运行，创建默认文件
    save_vessels(_DEFAULT_VESSELS)
    return list(_DEFAULT_VESSELS)


def save_vessels(vessels):
    """保存船舶列表到文件"""
    with open(VESSEL_FILE, "w", encoding="utf-8") as f:
        json.dump(vessels, f, ensure_ascii=False, indent=2)
    print(f"✅ 船舶列表已保存至 {VESSEL_FILE}")


def add_vessel(name, voyage, remark=""):
    """添加一艘船（支持备注）"""
    vessels = load_vessels()
    vessels = [v for v in vessels if not (v["name"] == name and v["voyage"] == voyage)]
    vessel = {"name": name, "voyage": voyage}
    if remark:
        vessel["remark"] = remark
    vessels.append(vessel)
    save_vessels(vessels)
    print(f"📌 已添加: {name} 航次 {voyage}" + (f" 备注: {remark}" if remark else ""))
    _git_commit_push(f"add vessel: {name} {voyage}")


def set_remark(name, remark):
    """修改船舶备注"""
    vessels = load_vessels()
    count = 0
    for v in vessels:
        if v["name"] == name:
            if remark:
                v["remark"] = remark
            else:
                v.pop("remark", None)
            count += 1
    if count:
        save_vessels(vessels)
        print(f"📝 已更新 [{name}] 的备注: {remark}")
        _git_commit_push(f"remark: {name} - {remark}")
    else:
        print(f"⚠️ 未找到名为 [{name}] 的船")


def remove_vessel(name):
    """按船名移除船舶"""
    vessels = load_vessels()
    before = len(vessels)
    vessels = [v for v in vessels if v["name"] != name]
    removed = before - len(vessels)
    if removed:
        save_vessels(vessels)
        print(f"🗑️ 已移除 {removed} 艘名为 [{name}] 的船")
        _git_commit_push(f"remove vessel: {name}")
    else:
        print(f"⚠️ 未找到名为 [{name}] 的船")


def list_vessels():
    """列出所有船舶"""
    vessels = load_vessels()
    if not vessels:
        print("📭 船舶列表为空")
        return
    print(f"📋 当前追踪 {len(vessels)} 艘船:")
    for i, v in enumerate(vessels, 1):
        remark = v.get("remark", "")
        print(f"  {i}. {v['name']} 航次 {v['voyage']}" + (f"  📝 {remark}" if remark else ""))


def _git_commit_push(msg):
    """自动提交并推送 git 变更（优先 gh CLI，其次 GitHub API + PAT）"""
    committed = False
    try:
        subprocess.run(["git", "add", VESSEL_FILE], capture_output=True, timeout=10)
        r = subprocess.run(["git", "commit", "-m", msg], capture_output=True, text=True, timeout=10)
        if r.returncode == 0:
            committed = True
    except:
        pass

    if not committed:
        print("   (无变更需提交)")
        return

    # 尝试 gh CLI 推送
    try:
        subprocess.run(["gh", "--version"], capture_output=True, timeout=5)
        result = subprocess.run(["git", "push"], capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            print("🚀 已推送至 GitHub")
            return
    except:
        pass

    # 尝试 GitHub API + PAT 推送
    if os.path.exists(GITHUB_PAT_FILE):
        with open(GITHUB_PAT_FILE, "r") as f:
            pat = f.read().strip()
        if pat:
            try:
                import base64
                with open(VESSEL_FILE, "rb") as f:
                    content_b64 = base64.b64encode(f.read()).decode()
                # 获取当前文件 SHA
                headers = {"Authorization": f"Bearer {pat}", "Accept": "application/vnd.github.v3+json"}
                resp = requests.get(
                    f"https://api.github.com/repos/{GITHUB_REPO}/contents/{VESSEL_FILE}",
                    headers=headers, timeout=10
                )
                sha = resp.json().get("sha") if resp.status_code == 200 else None
                # 更新文件
                data = {"message": msg, "content": content_b64}
                if sha:
                    data["sha"] = sha
                resp = requests.put(
                    f"https://api.github.com/repos/{GITHUB_REPO}/contents/{VESSEL_FILE}",
                    headers=headers, json=data, timeout=15
                )
                if resp.status_code in (200, 201):
                    print("🚀 已通过 API 推送至 GitHub")
                    return
            except:
                pass

    print("   ⚠️ 本地已保存，自动推送失败。手动执行: git push")

def get_webhook():
    """获取钉钉 Webhook URL（优先从 .webhook.txt 读取）"""
    if os.path.exists(WEBHOOK_FILE):
        try:
            with open(WEBHOOK_FILE, "r") as f:
                url = f.read().strip()
                if url:
                    return url
        except:
            pass
    return DINGTALK_WEBHOOK


def set_webhook(url):
    """设置钉钉 Webhook URL 并保存到文件"""
    with open(WEBHOOK_FILE, "w") as f:
        f.write(url.strip())
    print(f"✅ 钉钉 Webhook 已更新")
    _git_commit_push("update webhook")


# ==========================================
# Token 管理
# ==========================================

def load_token():
    """加载 Token：优先环境变量（GitHub Actions），其次本地文件"""
    token = os.environ.get("NPEDI_TOKEN")
    if token:
        return token
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None


def save_token(token):
    """保存 Token 到本地文件"""
    with open(TOKEN_FILE, "w", encoding="utf-8") as f:
        f.write(token)
    print(f"✅ Token 已保存至 {TOKEN_FILE}")


def grab_token_via_browser():
    """启动 Playwright 浏览器抓取 Token（仅本地运行有效）"""
    print("\n[系统提示] 正在启动浏览器环境以获取新 Token...")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("❌ 请先安装 Playwright: pip install playwright && playwright install chromium")
        return None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        token_container = []

        def on_request(request):
            try:
                for key, value in request.headers.items():
                    if key.lower() == "ediauthorization" and not token_container:
                        token_container.append(value)
                        print("🎉 成功拦截到底层请求，已捕获 Token！")
            except:
                pass

        context.on("request", on_request)
        print(f"👉 正在打开网页 {TOKEN_PAGE_URL}，请在浏览器中完成登录...")

        try:
            page.goto(TOKEN_PAGE_URL, timeout=90000, wait_until="domcontentloaded")
        except:
            pass

        for _ in range(300):
            if token_container:
                break
            page.wait_for_timeout(1000)

        if not token_container:
            print("\n❌ 抓取超时：未能获取到 Token。")
            browser.close()
            return None

        final_token = token_container[0]
        browser.close()
        return final_token

# ==========================================
# 数据查询核心
# ==========================================

def fetch_and_parse(vessel_name, voyage, current_token):
    """调用 API 并返回解析后的字典"""
    today = datetime.now()
    eta_begin = (today - timedelta(days=15)).strftime("%Y-%m-%d")
    eta_end = (today + timedelta(days=30)).strftime("%Y-%m-%d")

    params = {
        "vesselEnName": vessel_name, "voyage": voyage,
        "etaBegin": eta_begin, "etaEnd": eta_end,
        "page": "1", "pageSize": "50"
    }
    headers = {
        "Ediauthorization": current_token,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }

    try:
        response = requests.get(API_URL, params=params, headers=headers, timeout=10)
        if response.status_code == 401:
            return "EXPIRED"
        response.raise_for_status()
        data = response.json()

        # API 返回 code=401 也表示 Token 过期（但 HTTP 状态码是 200）
        if data.get('code') == 401:
            print(f"\n⚠️ [{vessel_name}] Token 已过期（API 返回 code=401）")
            return "EXPIRED"

        if data.get('code') != 200 or not data.get('data') or not data['data'].get('list'):
            print(f"\n⚠️ 未查询到 [{vessel_name}] 航次 [{voyage}] 的排期信息。")
            return "NO_DATA"

        item = data['data']['list'][0]

        parsed_data = {
            'update_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'cn_name': item.get('vesselCnName') or '-',
            'en_name': item.get('vesselEnName') or '-',
            'voyage': item.get('voyage') or '-',
            'terminal': item.get('terminal') or '-',
            'eta': item.get('eta') or '暂无',
            'etd': item.get('etd') or '暂无',
            'ata': item.get('ata') or '【尚未靠泊】',
            'atd': item.get('atd') or '【尚未离泊】',
            'ctn_start': item.get('ctnStartTime') or '-',
            'ctn_end': item.get('ctnEndTime') or '-',
            'port_close': item.get('portCloseTime') or '-',
            'custom_close': item.get('customCloseTime') or '-',
            'last_port': item.get('lastPortName') or '-',
            'next_port': item.get('nextPortName') or '-'
        }

        # 控制台打印
        print(f"\n📌 船名: {parsed_data['cn_name']} ({parsed_data['en_name']}) | 🔖 航次: {parsed_data['voyage']} | 🏗️ 码头: {parsed_data['terminal']}")
        print("-" * 45)
        print(f"⏳ 计划靠泊 (ETA): {parsed_data['eta']} | 计划离泊 (ETD): {parsed_data['etd']}")
        print(f"⚓ 实际靠泊 (ATA): {parsed_data['ata']} | 实际离泊 (ATD): {parsed_data['atd']}")
        print(f"📦 进箱时间: {parsed_data['ctn_start']} 至 {parsed_data['ctn_end']}")
        print(f"🛑 码头截单: {parsed_data['port_close']} | 海关截关: {parsed_data['custom_close']}")
        print(f"🗺️ 航线轨迹: 上一港 [{parsed_data['last_port']}] ➔ 下一港 [{parsed_data['next_port']}]\n")

        return parsed_data

    except Exception as e:
        print(f"\n❌ 网络请求异常: {e}")
        return "ERROR"

# ==========================================
# 钉钉推送 & Excel 写入
# ==========================================

def send_dingtalk_msg(data_list):
    """将数据排版为 Markdown 并推送到钉钉（含备注）"""
    if not data_list:
        return

    # 加载船舶列表以获取备注
    vessels = load_vessels()
    remark_map = {}
    for v in vessels:
        r = v.get("remark", "")
        if r:
            remark_map[f"{v['name']}_{v['voyage']}"] = r

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    content = f"### 🚢 船舶排期自动更新\n*{current_time}*\n\n---\n\n"

    for item in data_list:
        key = f"{item['en_name']}_{item['voyage']}"
        remark = remark_map.get(key, "")
        if remark:
            content += f"# 📝 {remark}\n\n"
        name_line = f"**📌 船名**: {item['cn_name']} ({item['en_name']})"
        name_line += f" | **🔖 航次**: {item['voyage']} | **🏗️ 码头**: {item['terminal']}\n\n"
        content += name_line
        content += f"**⏳ 计划靠泊 (ETA)**: <font color=#008000>{item['eta']}</font> | **计划离泊 (ETD)**: {item['etd']}\n\n"

        # 状态变化点：有实际时间时绿色加粗标注
        def _green_bold(v):
            return f"<font color=#008000><b>{v}</b></font>" if v and "尚未" not in str(v) and v not in ("暂无", "-", "") else v

        content += f"**⚓ 实际靠泊 (ATA)**: {_green_bold(item['ata'])} | **实际离泊 (ATD)**: {_green_bold(item['atd'])}\n\n"
        content += f"**📦 进箱时间**: {item['ctn_start']} 至 {item['ctn_end']}\n\n"
        content += f"**🛑 码头截单**: <font color=#FF0000>{item['port_close']}</font> | **海关截关**: {item['custom_close']}\n\n"
        content += f"**🗺️ 航线轨迹**: 上一港 [{item['last_port']}] ➔ 下一港 [{item['next_port']}]\n\n"
        content += "---\n\n"

    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": "🚢 船舶动态更新",
            "text": content
        }
    }

    try:
        response = requests.post(get_webhook(), json=payload)
        if response.status_code == 200:
            print("📣 钉钉消息推送成功！")
        else:
            print(f"⚠️ 钉钉推送失败: {response.text}")
    except Exception as e:
        print(f"⚠️ 钉钉推送出现异常: {e}")


def _was_last_run_also_failed():
    """检查 GitHub Actions 上一次运行是否也失败了（避免重复发钉钉告警）"""
    if "GITHUB_ACTIONS" not in os.environ:
        return False  # 本地运行，不检查

    gh_token = os.environ.get("GITHUB_TOKEN")
    repo = os.environ.get("GITHUB_REPOSITORY")
    run_id = os.environ.get("GITHUB_RUN_ID")
    workflow_name = os.environ.get("GITHUB_WORKFLOW")
    if not all([gh_token, repo, run_id, workflow_name]):
        return False

    try:
        # 查询当前 workflow 的最近 2 次完成运行
        resp = requests.get(
            f"https://api.github.com/repos/{repo}/actions/runs",
            headers={
                "Authorization": f"Bearer {gh_token}",
                "Accept": "application/vnd.github.v3+json"
            },
            params={
                "status": "completed",
                "per_page": 2
            },
            timeout=10
        )
        if resp.status_code != 200:
            return False

        runs = resp.json().get("workflow_runs", [])
        if len(runs) < 2:
            return False  # 只有一次运行，没有"上一次"

        # 跳过当前 run，检查上一次
        for r in runs:
            if str(r["id"]) == str(run_id):
                continue
            return r.get("conclusion") == "failure"

    except Exception:
        pass
    return False


def send_email_alert(subject, body):
    """通过 QQ 邮箱 SMTP 发送告警邮件"""
    import smtplib
    from email.mime.text import MIMEText

    smtp_user = "1642425527@qq.com"
    smtp_pass = os.environ.get("QQ_SMTP_PASS")
    to_email = "1642425527@qq.com"

    if not smtp_pass:
        # 本地运行：尝试从 .smtp.txt 读取
        smtp_file = _path(".smtp.txt")
        if os.path.exists(smtp_file):
            with open(smtp_file, "r") as f:
                lines = [l.strip() for l in f.readlines() if l.strip()]
                if len(lines) >= 2:
                    smtp_user, smtp_pass = lines[0], lines[1]

    if not smtp_pass:
        print("⚠️ 未设置 QQ_SMTP_PASS 环境变量/找不到 .smtp.txt，跳过邮件告警。")
        return

    # 去重检查
    if "GITHUB_ACTIONS" in os.environ:
        if _was_last_run_also_failed():
            print(f"   ⏭️ 跳过重复邮件告警（上一次运行也已失败）")
            return

    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = smtp_user
        msg["To"] = to_email

        with smtplib.SMTP_SSL("smtp.qq.com", 465, timeout=15) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        print(f"📧 告警邮件已发送至 {to_email}: {subject}")
    except Exception as e:
        print(f"⚠️ 邮件发送失败: {e}")


# ==========================================
# 数据变化检测（避免重复推送）
# ==========================================

_IMPORTANT_FIELDS = ["eta", "etd", "ata", "atd", "ctn_start", "ctn_end",
                     "port_close", "custom_close", "terminal"]


def load_previous_state():
    """加载上次运行的船舶状态"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_current_state(data_list):
    """保存本次运行的船舶状态"""
    state = {}
    for item in data_list:
        key = f"{item['en_name']}_{item['voyage']}"
        state[key] = {f: item.get(f, "") for f in _IMPORTANT_FIELDS}
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    print(f"💾 已保存当前船态至 {STATE_FILE}")


def has_data_changed(data_list):
    """与上次数据对比 + 超期检测，返回是否有变化和一个变更摘要"""
    previous = load_previous_state()

    if not previous:
        return True, "🆕 首次运行，推送初始数据"

    changes = []
    for item in data_list:
        key = f"{item['en_name']}_{item['voyage']}"
        old = previous.get(key)
        if not old:
            changes.append(f"📌 {item['cn_name']} ({item['en_name']}) 新出现")
            continue
        for field in _IMPORTANT_FIELDS:
            new_val = item.get(field, "")
            old_val = old.get(field, "")
            if new_val != old_val:
                field_names = {
                    "eta": "计划靠泊(ETA)", "etd": "计划离泊(ETD)",
                    "ata": "实际靠泊(ATA)", "atd": "实际离泊(ATD)",
                    "ctn_start": "进箱开始", "ctn_end": "进箱结束",
                    "port_close": "码头截单", "custom_close": "海关截关",
                    "terminal": "码头"
                }
                name = field_names.get(field, field)
                changes.append(f"🔁 {item['cn_name']} {name}: {old_val} → {new_val}")

    # ⏰ 超期检测：ETA已过但未靠泊 / ETD已过但未离泊
    for item in data_list:
        now = datetime.now()
        name = item['cn_name']

        # ETA 已过但未靠泊
        eta_str = item.get("eta", "")
        ata_str = item.get("ata", "")
        if eta_str and eta_str != "暂无" and ("尚未" in ata_str or not ata_str or ata_str in ("暂无", "-", "")):
            try:
                eta_time = datetime.strptime(eta_str, "%Y-%m-%d %H:%M:%S")
                if eta_time < now:
                    changes.append(f"⏰ {name} ETA {eta_str} 已过，仍未靠泊！")
            except:
                pass

        # ETD 已过但未离泊
        etd_str = item.get("etd", "")
        atd_str = item.get("atd", "")
        if etd_str and etd_str != "暂无" and ("尚未" in atd_str or not atd_str or atd_str in ("暂无", "-", "")):
            try:
                etd_time = datetime.strptime(etd_str, "%Y-%m-%d %H:%M:%S")
                if etd_time < now:
                    changes.append(f"⏰ {name} ETD {etd_str} 已过，仍未离泊！")
            except:
                pass

    if changes:
        return True, "数据变更:\n" + "\n".join(changes)
    return False, "数据无变化"


def save_to_excel(data_list):
    """将数据写入 Excel 并进行简单排版"""
    if not data_list:
        return

    file_exists = os.path.exists(EXCEL_FILE)

    try:
        if file_exists:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "船舶排期"
            headers = ["更新时间", "船名 (英文)", "航次", "码头", "计划靠泊(ETA)", "计划离泊(ETD)",
                       "实际靠泊(ATA)", "实际离泊(ATD)", "进箱开始", "进箱结束",
                       "码头截单", "海关截关", "航线预告"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['M'].width = 30

        for item in data_list:
            row = [
                item['update_time'],
                f"{item['cn_name']} ({item['en_name']})",
                item['voyage'],
                item['terminal'],
                item['eta'], item['etd'],
                item['ata'], item['atd'],
                item['ctn_start'], item['ctn_end'],
                item['port_close'], item['custom_close'],
                f"上港:{item['last_port']} -> 下港:{item['next_port']}"
            ]
            ws.append(row)

        wb.save(EXCEL_FILE)
        print(f"📁 数据已成功保存至 Excel 文档: {EXCEL_FILE}")

    except PermissionError:
        print(f"❌ 写入 Excel 失败！请确认 `{EXCEL_FILE}` 没有被其他软件(如 Office) 打开。")
    except Exception as e:
        print(f"❌ 保存 Excel 发生异常: {e}")

# ==========================================
# Token 推送至 GitHub
# ==========================================

GITHUB_REPO = "a1642425527-ship-it/ship-tracker"
GITHUB_PAT_FILE = _path("GITHUB_PAT.txt")


def push_token_to_github(new_token):
    """将新 Token 推送到 GitHub Secrets（自动选择可用方式）"""
    print("\n🔄 正在将新 Token 推送到 GitHub Secrets...")

    # 方法1: gh CLI
    if _try_push_via_gh(new_token):
        return True

    # 方法2: GitHub API + PAT
    if _try_push_via_api(new_token):
        return True

    # 都失败，打印手动操作指引
    print(f"\n💡 请手动更新 GitHub Secret:")
    print(f"   ① 打开 https://github.com/{GITHUB_REPO}/settings/secrets/actions")
    print(f"   ② 点击 NPEDI_TOKEN 右边的编辑按钮")
    print(f"   ③ 粘贴以下 Token 并保存:")
    print(f"      {new_token}")
    return False


def _try_push_via_gh(new_token):
    """通过 gh CLI 推送"""
    try:
        subprocess.run(["gh", "--version"], capture_output=True, check=True, timeout=5)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False

    try:
        result = subprocess.run(
            ["gh", "auth", "status"], capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return False

        result = subprocess.run(
            ["gh", "secret", "set", "NPEDI_TOKEN", "--repo", GITHUB_REPO],
            input=new_token, capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            print("✅ Token 已通过 gh CLI 推送至 GitHub Secrets！")
            return True
    except Exception:
        pass
    return False


def _try_push_via_api(new_token):
    """通过 GitHub API + PAT 推送"""
    if not os.path.exists(GITHUB_PAT_FILE):
        print(f"   ⚠️ 未找到 {GITHUB_PAT_FILE}，跳过 API 方式。")
        print(f"   💡 创建 {GITHUB_PAT_FILE} 后即可自动推送（见 README）")
        return False

    with open(GITHUB_PAT_FILE, "r", encoding="utf-8") as f:
        pat = f.read().strip()

    if not pat:
        return False

    headers = {
        "Authorization": f"Bearer {pat}",
        "Accept": "application/vnd.github.v3+json"
    }

    try:
        # 获取公钥
        print("   📡 正在获取 GitHub Secrets 公钥...")
        resp = requests.get(
            f"https://api.github.com/repos/{GITHUB_REPO}/actions/secrets/public-key",
            headers=headers, timeout=15
        )
        if resp.status_code != 200:
            print(f"   ⚠️ 获取公钥失败 (HTTP {resp.status_code})")
            return False

        key_data = resp.json()
        key_id = key_data["key_id"]
        public_key_b64 = key_data["key"]

        # 加密 Token（使用 libsodium crypto_box_seal）
        try:
            import nacl.bindings
            import base64
            pub_key_bytes = base64.b64decode(public_key_b64)
            encrypted = nacl.bindings.crypto_box_seal(new_token.encode(), pub_key_bytes)
            encrypted_b64 = base64.b64encode(encrypted).decode()
        except ImportError:
            print("   ⚠️ 需要安装 PyNaCl 才能加密: pip install pynacl")
            return False

        # 写入 Secret
        resp = requests.put(
            f"https://api.github.com/repos/{GITHUB_REPO}/actions/secrets/NPEDI_TOKEN",
            headers=headers,
            json={"encrypted_value": encrypted_b64, "key_id": key_id},
            timeout=15
        )
        if resp.status_code in (201, 204):
            print("✅ Token 已通过 API 推送至 GitHub Secrets！（下次 Action 自动使用）")
            return True
        else:
            print(f"   ⚠️ 设置 Secret 失败: {resp.text}")
            return False

    except requests.exceptions.RequestException as e:
        print(f"   ⚠️ API 请求失败: {e}")
        return False

# ==========================================
# 执行器
# ==========================================

def _is_quiet_hours():
    """判断当前是否在静默时段（CST 23:00 - 07:00）"""
    if "GITHUB_ACTIONS" not in os.environ:
        return False  # 本地运行不限制
    cst_hour = (datetime.utcnow() + timedelta(hours=8)).hour
    return cst_hour >= 23 or cst_hour < 7


def run_once(force_push=False):
    """单次查询（GitHub Actions + 本地单次运行）
       force_push=True 时忽略变化检测，强制推钉钉"""
    # 优先从环境变量读取强制推送标志（GitHub Actions 手动触发时传入）
    if os.environ.get("FORCE_PUSH") == "true":
        force_push = True
    print(f"\n🚀 开始查询船舶动态... ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})")

    api_token = load_token()
    if not api_token:
        send_email_alert(
            "⛔ 船舶跟踪系统 - Token 未配置",
            "船舶动态跟踪系统：未找到有效的 NPEDI Token\n\n"
            "解决方法：\n"
            "1. 在 Windows 上进入 ship-tracker 目录\n"
            "2. 运行 python ship_tracker.py --grab-token 抓取新 Token\n"
            "3. 脚本会自动推送到 GitHub Secrets"
        )
        print("❌ 未找到 Token！")
        print("💡 请先本地运行: python ship_tracker.py --grab-token")
        print("   或者设置环境变量: export NPEDI_TOKEN=你的Token")
        sys.exit(1)

    results_to_save = []
    vessels = load_vessels()

    for index, vessel in enumerate(vessels):
        target_name = vessel.get("name")
        target_voyage = vessel.get("voyage")

        while True:
            result = fetch_and_parse(target_name, target_voyage, api_token)

            if result == "EXPIRED":
                send_email_alert(
                    "⛔ 船舶跟踪系统 - NPEDI Token 已过期",
                    f"船舶动态跟踪系统：NPEDI Token 已失效\n\n"
                    f"受影响船舶：{target_name} 航次 {target_voyage}\n\n"
                    f"解决方法：\n"
                    f"1. 在 Windows 上进入 ship-tracker 目录\n"
                    f"2. 运行 python ship_tracker.py --grab-token 抓取新 Token\n"
                    f"3. 脚本会自动推送到 GitHub Secrets"
                )
                print("⚠️ Token 已过期！请重新抓取。")
                print(f"💡 本地运行: python ship_tracker.py --grab-token")
                sys.exit(1)
            elif isinstance(result, dict):
                results_to_save.append(result)
                break
            else:
                break

        if index < len(vessels) - 1:
            time.sleep(round(random.uniform(1.5, 3.5), 2))

    if results_to_save:
        changed, reason = has_data_changed(results_to_save)
        save_current_state(results_to_save)
        save_to_excel(results_to_save)

        if changed or force_push:
            if _is_quiet_hours() and not force_push:
                print(f"\n🌙 静默时段（CST 23:00-07:00），跳过钉钉推送")
            else:
                send_dingtalk_msg(results_to_save)
                print(f"\n📣 已推送钉钉{'（强制推送）' if force_push else ''}")
                print(f"   {reason}")
        else:
            print(f"\n⏭️ 数据无变化，跳过钉钉推送")

        print(f"✅ 本轮查询完成！")
        return True

    print(f"\n⚠️ 未查询到任何数据。")
    return False


def run_daemon():
    """本地守护进程模式（每2小时自动查询）"""
    import threading

    force_run_event = threading.Event()

    def input_listener():
        while True:
            try:
                input()
                print("\n⚡ 检测到回车指令，立即触发手动查询！")
                force_run_event.set()
            except:
                break

    print(f"🌟 港口船舶自动化跟踪系统已启动！")
    print(f"⏲️ 当前设置：每 {UPDATE_INTERVAL_HOURS} 小时自动更新一次数据至 Excel 并推送钉钉。")
    print(f"💡 随时按【回车键 (Enter)】即可立即强制拉取最新数据。")
    print("=" * 60)

    listener_thread = threading.Thread(target=input_listener, daemon=True)
    listener_thread.start()

    while True:
        run_once()

        next_run_time = datetime.now() + timedelta(hours=UPDATE_INTERVAL_HOURS)
        print(f"⏳ 下一次自动执行时间为: {next_run_time.strftime('%Y-%m-%d %H:%M:%S')}")

        force_run_event.clear()
        while datetime.now() < next_run_time:
            if force_run_event.is_set():
                break
            time.sleep(1)

# ==========================================
# 入口
# ==========================================

if __name__ == "__main__":
    if "--grab-token" in sys.argv:
        new_token = grab_token_via_browser()
        if new_token:
            save_token(new_token)
            print(f"\n🔑 新 Token: {new_token[:20]}...{new_token[-10:]}")
            push_token_to_github(new_token)

    elif "--daemon" in sys.argv:
        run_daemon()

    elif "--push" in sys.argv:
        print("🚀 正在触发 GitHub Actions 远程查询...")
        if os.path.exists(GITHUB_PAT_FILE):
            with open(GITHUB_PAT_FILE, "r") as f:
                pat = f.read().strip()
            if pat:
                try:
                    resp = requests.post(
                        f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/ship-track.yml/dispatches",
                        headers={
                            "Authorization": f"Bearer {pat}",
                            "Accept": "application/vnd.github.v3+json"
                        },
                        json={"ref": "main", "inputs": {"force_push": "true"}},
                        timeout=15
                    )
                    if resp.status_code == 204:
                        print("✅ 已触发 GitHub Action（强制推送模式），稍后查看钉钉")
                    else:
                        print(f"⚠️ 触发失败 (HTTP {resp.status_code})，尝试本地运行...")
                        run_once(force_push=True)
                except Exception as e:
                    print(f"⚠️ 触发远程失败: {e}，尝试本地运行...")
                    run_once(force_push=True)
            else:
                print("⚠️ GITHUB_PAT.txt 为空，尝试本地运行...")
                run_once(force_push=True)
        else:
            print("⚠️ 未找到 GITHUB_PAT.txt，尝试本地运行...")
            run_once(force_push=True)

    elif "--add" in sys.argv:
        if len(sys.argv) >= 5:
            add_vessel(sys.argv[2], sys.argv[3], sys.argv[4])
        elif len(sys.argv) >= 4:
            add_vessel(sys.argv[2], sys.argv[3])
        else:
            print("⚠️ 用法: python ship_tracker.py --add \"船名\" \"航次\" [\"备注\"]")

    elif "--remark" in sys.argv:
        if len(sys.argv) >= 4:
            set_remark(sys.argv[2], sys.argv[3])
        else:
            print("⚠️ 用法: python ship_tracker.py --remark \"船名\" \"备注内容\"")

    elif "--remove" in sys.argv:
        if len(sys.argv) >= 3:
            remove_vessel(sys.argv[2])
        else:
            print("⚠️ 用法: python ship_tracker.py --remove \"船名\"")

    elif "--list" in sys.argv:
        list_vessels()

    elif "--enable" in sys.argv:
        if os.path.exists(DISABLED_FILE):
            os.remove(DISABLED_FILE)
            print("✅ 查询已启用")
            _git_commit_push("enable query")
        else:
            print("✅ 查询已经是启用状态")

    elif "--disable" in sys.argv:
        if not os.path.exists(DISABLED_FILE):
            open(DISABLED_FILE, "w").close()
            print("⏸️ 查询已禁用（GitHub Action 将跳过查询）")
            _git_commit_push("disable query")
        else:
            print("⏸️ 查询已经是禁用状态")

    elif "--webhook" in sys.argv:
        if len(sys.argv) >= 3:
            set_webhook(sys.argv[2])
        else:
            print(f"📡 当前 Webhook: {get_webhook()[:50]}...")
            print("⚠️ 用法: python ship_tracker.py --webhook \"新URL\"")

    elif "--status" in sys.argv:
        print(f"📋 船舶跟踪系统状态")
        print(f"{'='*40}")
        print(f"⏰ 查询间隔: 每 {UPDATE_INTERVAL_HOURS} 小时")
        print(f"{'🔴 已禁用' if os.path.exists(DISABLED_FILE) else '🟢 运行中'}")
        list_vessels()
        token = load_token()
        print(f"🔑 Token: {'✅ 已存在' if token else '❌ 未配置'}")
        if os.path.exists(STATE_FILE):
            print(f"💾 上次状态缓存: 存在")
        else:
            print(f"💾 上次状态缓存: 无")

    else:
        # 默认：单次查询模式（GitHub Actions）
        if os.path.exists(DISABLED_FILE):
            print("⏸️ 查询已禁用（.disabled 文件存在），跳过本次查询。")
            sys.exit(0)
        run_once()
